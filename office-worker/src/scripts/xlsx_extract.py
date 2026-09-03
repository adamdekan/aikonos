#!/usr/bin/env python3
"""Fixed helper (NOT model-authored): xlsx.extract — sheet/range -> CSV/JSON
summary via openpyxl (cell/range reads) + pandas (serialization). Declarative
params only, no script execution.

argv: <input.xlsx> <params.json> <output.json>

params.json: {"sheet": str (optional, default first sheet), "range": str
(optional, e.g. "A1:C10"), "format": "csv"|"json" (default "csv")}
"""
import json
import sys

import openpyxl
import pandas as pd
from openpyxl.utils.cell import range_boundaries


def main():
    input_path, params_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]
    with open(params_path, "r", encoding="utf-8") as f:
        params = json.load(f)

    sheet_name = params.get("sheet")
    range_str = params.get("range")
    fmt = params.get("format", "csv")

    # data_only=True reads cached formula results, not formula source text —
    # correct for an extract (recalculating live is xlsx.recalc's job, CP3).
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    if range_str:
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        rows = list(
            ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
        )
    else:
        rows = list(ws.iter_rows(values_only=True))

    df = pd.DataFrame(rows)
    content = df.to_json(orient="values") if fmt == "json" else df.to_csv(index=False, header=False)

    result = {"sheet": ws.title, "rows": len(rows), "format": fmt, "content": content}
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f)


if __name__ == "__main__":
    main()
